"""
LiberStudy Discovery - Reddit Pain Point Scraper
自动搜索 Reddit 帖子并提取课堂笔记/AI工具相关痛点数据，输出带有话题分类的 Excel 文件。

使用方式：
    pip install requests openpyxl
    python reddit_scraper.py

输出：与脚本同目录下的 reddit_data.xlsx
"""

import requests
import time
import json
from datetime import datetime, timezone
from pathlib import Path

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Alignment, PatternFill
except ImportError:
    print("请先安装 openpyxl: pip install openpyxl")
    raise SystemExit(1)

# ============================================================
# 配置区 - 可按需调整
# ============================================================

SUBREDDITS = [
    "StudyTips",
    "college",
    "productivity",
    "NotebookLM",
    "artificial",
    "GradSchool",
]

# 话题分类：(话题索引, 话题名称, 关键词列表)
# 注意：已修复这里的中文引号为英文引号，否则会报 SyntaxError
# 精简版话题分类：从 27 个关键词去重/提纯到 13 个高价值词
TOPICS = [
    (1, "痛点与挣扎 (Pain Points)", [
        "lecture recording too long",  # 涵盖了听录音、复习耗时的痛点
        "struggle class notes",        # 缩短词组，提升命中率
        "hate lecture slides",         # 极度厌恶看PPT的情绪
        "overwhelmed lecture",         # 知识点过载的无力感
    ]),
    (2, "竞品吐槽与平替 (Alternatives)", [
        "Otter.ai alternative",        # 寻找替代品是极强的精准需求
        "Otter.ai sucks",              # 直接抓取竞品的核心缺陷
        "better than NotebookLM",      # 寻找更优解
        "tired of manual notes",       # 用户渴望自动化的前置情绪
    ]),
    (3, "诉求与土办法 (Workarounds)", [
        "speed up lecture review",     # 合并了所有关于“复习提效”的需求
        "app summarizes lectures",     # 明确的工具诉求
        "workaround lecture recording", # 抓取用户的“土办法”
    ]),
    (4, "高意向求助 (High Intent)", [
        "best AI class notes",         # 去掉介词，抓取求推荐贴
        "transcribe lecture AI",       # 核心场景的最精简表达
    ]),
]

# 自动生成：关键词 → (话题索引, 话题名称) 的映射
KEYWORD_TO_TOPIC = {}
KEYWORDS = []
for topic_idx, topic_name, kws in TOPICS:
    for kw in kws:
        KEYWORD_TO_TOPIC[kw] = (topic_idx, topic_name)
        KEYWORDS.append(kw)

POSTS_PER_SEARCH = 7          # 每次搜索取 top N 个帖子
TOP_COMMENTS = 7              # 每个帖子取 top N 条评论
REQUEST_DELAY = 1.4            # 请求间隔（秒），避免 429
OUTPUT_DIR = Path(__file__).parent
OUTPUT_FILE = OUTPUT_DIR / "reddit_data.xlsx"

HEADERS = {
    "User-Agent": "LiberStudy-Discovery/1.1 (research; contact: liberstudy@example.com)"
}

COOLDOWN_THRESHOLD = 2            # 连续 429 达到此数时触发长等待
COOLDOWN_SECONDS = 120            # 长等待时间（秒）
MAX_CONSECUTIVE_RATE_LIMITS = 3   # 连续被限流次数上限（终止爬取）
_consecutive_429_count = 0        # 全局连续 429 计数器


class RateLimitAbort(Exception):
    """连续多次被限流，提前终止爬取。"""
    def __init__(self, message: str):
        self.message = message
        super().__init__(message)


def _check_rate_limit(resp, context: str) -> bool:
    """检查 429 并更新计数器。返回 True 表示需要重试。"""
    global _consecutive_429_count
    if resp.status_code == 429:
        _consecutive_429_count += 1
        print(f"  [429] 被限流 (连续第 {_consecutive_429_count} 次)")
        if _consecutive_429_count >= MAX_CONSECUTIVE_RATE_LIMITS:
            raise RateLimitAbort(context)
        if _consecutive_429_count >= COOLDOWN_THRESHOLD:
            print(f"  ⏸️  连续 {COOLDOWN_THRESHOLD} 次被限流，暂停 {COOLDOWN_SECONDS} 秒后继续...")
            time.sleep(COOLDOWN_SECONDS)
        else:
            time.sleep(10)
        return True  # 需要重试
    else:
        _consecutive_429_count = 0  # 成功请求，重置计数
        return False


# ============================================================
# 核心函数
# ============================================================

def search_subreddit(subreddit: str, query: str, limit: int = 10) -> list[dict]:
    """搜索 subreddit，返回帖子基本信息列表。"""
    url = f"https://www.reddit.com/r/{subreddit}/search.json"
    params = {
        "q": query,
        "sort": "relevance",
        "t": "year",           # 最近一年
        "restrict_sr": "on",   # 限定在该 subreddit
        "limit": limit,
    }
    try:
        resp = requests.get(url, headers=HEADERS, params=params, timeout=15)
        if _check_rate_limit(resp, f"搜索 r/{subreddit} '{query}'"):
            resp = requests.get(url, headers=HEADERS, params=params, timeout=15)
            _check_rate_limit(resp, f"搜索 r/{subreddit} '{query}' (重试)")
        resp.raise_for_status()
        data = resp.json()
        posts = []
        for child in data.get("data", {}).get("children", []):
            p = child.get("data", {})
            posts.append({
                "subreddit": p.get("subreddit", ""),
                "title": p.get("title", ""),
                "selftext": (p.get("selftext", "") or "")[:2000],  # 截断过长正文
                "score": p.get("score", 0),
                "upvote_ratio": p.get("upvote_ratio", 0),
                "num_comments": p.get("num_comments", 0),
                "created_utc": p.get("created_utc", 0),
                "permalink": p.get("permalink", ""),
                "id": p.get("id", ""),
            })
        return posts
    except Exception as e:
        print(f"  [ERROR] 搜索 r/{subreddit} '{query}' 失败: {e}")
        return []

def fetch_top_comments(permalink: str, top_n: int = 3) -> list[str]:
    """获取帖子的 top N 条评论。"""
    url = f"https://www.reddit.com{permalink}.json"
    try:
        resp = requests.get(url, headers=HEADERS, timeout=15)
        if _check_rate_limit(resp, f"获取评论 {permalink}"):
            resp = requests.get(url, headers=HEADERS, timeout=15)
            _check_rate_limit(resp, f"获取评论 {permalink} (重试)")
        resp.raise_for_status()
        data = resp.json()
        if not isinstance(data, list) or len(data) < 2:
            return []
        comments_data = data[1].get("data", {}).get("children", [])
        # 过滤掉 "more" 类型和机器人
        comments = []
        for c in comments_data:
            if c.get("kind") != "t1":
                continue
            cd = c.get("data", {})
            body = cd.get("body", "")
            score = cd.get("score", 0)
            if body and len(body) > 10:  # 过滤太短的评论
                comments.append({"body": body[:1000], "score": score})
        # 按 score 排序取 top N
        comments.sort(key=lambda x: x["score"], reverse=True)
        return [c["body"] for c in comments[:top_n]]
    except Exception as e:
        print(f"  [ERROR] 获取评论失败 {permalink}: {e}")
        return []

def utc_to_readable(utc_ts: float) -> str:
    """Unix 时间戳转可读日期。"""
    if not utc_ts:
        return ""
    return datetime.fromtimestamp(utc_ts, tz=timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

def _build_columns() -> list[tuple[str, int]]:
    """根据 TOP_COMMENTS 动态生成列定义。"""
    cols = [
        ("话题索引", 10),
        ("话题分类", 20),
        ("搜索关键词", 25),
        ("Subreddit", 15),
        ("帖子标题", 50),
        ("帖子正文", 60),
        ("帖子得分", 10),
        ("点赞率", 10),
        ("评论总数", 10),
        ("发帖时间", 22),
    ]
    for i in range(1, TOP_COMMENTS + 1):
        cols.append((f"Top评论{i}", 60))
    cols.append(("帖子链接", 50))
    return cols


def load_existing_data(filepath: Path) -> dict[str, dict]:
    """从已有 xlsx 中读取数据，返回 {post_id: row_dict} 映射。"""
    if not filepath.exists():
        return {}
    try:
        wb = load_workbook(filepath, read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(min_row=1, values_only=True)
        header = next(rows_iter, None)
        if not header:
            wb.close()
            return {}

        # 找到帖子链接列的索引
        link_col = None
        comment_cols = []
        for idx, h in enumerate(header):
            if h == "帖子链接":
                link_col = idx
            if h and str(h).startswith("Top评论"):
                comment_cols.append(idx)

        if link_col is None:
            wb.close()
            return {}

        existing = {}
        for row_values in rows_iter:
            if row_values is None:
                continue
            link = row_values[link_col] if link_col < len(row_values) else ""
            if not link:
                continue
            # 从链接中提取 post id（链接格式: https://www.reddit.com/r/.../comments/ID/...）
            post_id = _extract_id_from_link(link)
            if not post_id:
                continue

            comments = []
            for ci in comment_cols:
                val = row_values[ci] if ci < len(row_values) else ""
                if val:
                    comments.append(str(val))

            existing[post_id] = {
                "topic_idx": row_values[0] if len(row_values) > 0 else "",
                "topic_name": row_values[1] if len(row_values) > 1 else "",
                "keyword": row_values[2] if len(row_values) > 2 else "",
                "subreddit": row_values[3] if len(row_values) > 3 else "",
                "title": row_values[4] if len(row_values) > 4 else "",
                "selftext": row_values[5] if len(row_values) > 5 else "",
                "score": row_values[6] if len(row_values) > 6 else 0,
                "upvote_ratio": row_values[7] if len(row_values) > 7 else 0,
                "num_comments": row_values[8] if len(row_values) > 8 else 0,
                "created_readable": row_values[9] if len(row_values) > 9 else "",
                "top_comments": comments,
                "permalink": link.replace("https://www.reddit.com", "") if link else "",
            }
        wb.close()
        print(f"📂 已加载旧数据: {len(existing)} 条帖子")
        return existing
    except Exception as e:
        print(f"  [WARN] 读取已有 xlsx 失败，将重新创建: {e}")
        return {}


def _extract_id_from_link(link: str) -> str:
    """从 Reddit 链接中提取帖子 ID。"""
    # https://www.reddit.com/r/xxx/comments/ABC123/...
    parts = link.split("/")
    try:
        idx = parts.index("comments")
        return parts[idx + 1] if idx + 1 < len(parts) else ""
    except ValueError:
        return ""


def merge_rows(existing: dict[str, dict], new_rows: list[dict]) -> list[dict]:
    """合并已有数据和新数据。重复帖子：补充评论；新帖子：直接添加。"""
    merged = dict(existing)  # 浅拷贝
    new_count = 0
    updated_count = 0

    for row in new_rows:
        permalink = row.get("permalink", "")
        post_id = _extract_id_from_link(f"https://www.reddit.com{permalink}")
        if not post_id:
            post_id = row.get("id", permalink)

        if post_id in merged:
            # 已有帖子：用新评论补充旧评论
            old_comments = merged[post_id].get("top_comments", [])
            new_comments = row.get("top_comments", [])
            # 去重合并（保留顺序，旧的在前，新的补在后面）
            seen_comments = set(old_comments)
            combined = list(old_comments)
            for c in new_comments:
                if c not in seen_comments:
                    combined.append(c)
                    seen_comments.add(c)
            if len(combined) > len(old_comments):
                updated_count += 1
            merged[post_id]["top_comments"] = combined[:TOP_COMMENTS]
        else:
            new_count += 1
            merged[post_id] = row

    print(f"   合并结果: {new_count} 条新帖子, {updated_count} 条帖子评论被补充")
    return list(merged.values())


def write_excel(rows: list[dict], filepath: Path):
    """将数据写入 Excel 文件。"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Reddit Data"

    columns = _build_columns()

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, (col_name, col_width) in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = col_width

    # 数据行写入
    for row_idx, row in enumerate(rows, 2):
        comments = row.get("top_comments", [])

        ws.cell(row=row_idx, column=1, value=row.get("topic_idx", ""))
        ws.cell(row=row_idx, column=2, value=row.get("topic_name", ""))
        ws.cell(row=row_idx, column=3, value=row.get("keyword", ""))
        ws.cell(row=row_idx, column=4, value=row.get("subreddit", ""))
        ws.cell(row=row_idx, column=5, value=row.get("title", ""))
        ws.cell(row=row_idx, column=6, value=row.get("selftext", ""))
        ws.cell(row=row_idx, column=7, value=row.get("score", 0))
        ws.cell(row=row_idx, column=8, value=row.get("upvote_ratio", 0))
        ws.cell(row=row_idx, column=9, value=row.get("num_comments", 0))
        ws.cell(row=row_idx, column=10, value=row.get("created_readable", ""))
        # 动态写入评论列
        for i in range(TOP_COMMENTS):
            ws.cell(row=row_idx, column=11 + i, value=comments[i] if i < len(comments) else "")
        # 帖子链接在评论列之后
        ws.cell(row=row_idx, column=11 + TOP_COMMENTS, value=f"https://www.reddit.com{row.get('permalink', '')}")

        # 动态处理所有列的自动换行
        for col in range(1, len(columns) + 1):
            ws.cell(row=row_idx, column=col).alignment = Alignment(wrap_text=True, vertical="top")

    # 冻结首行
    ws.freeze_panes = "A2"

    filepath.parent.mkdir(parents=True, exist_ok=True)
    wb.save(filepath)
    print(f"\n✅ Excel 已保存: {filepath}")
    print(f"   共 {len(rows)} 条数据")

# ============================================================
# 主流程
# ============================================================

def main():
    # 加载已有数据（如果 xlsx 存在）
    existing_data = load_existing_data(OUTPUT_FILE)

    all_rows = []

    total_searches = len(SUBREDDITS) * len(KEYWORDS)
    current = 0
    abort_message = None

    try:
        for keyword in KEYWORDS:
            # 提取当前 keyword 对应的索引和话题名称
            topic_idx, topic_name = KEYWORD_TO_TOPIC[keyword]

            for subreddit in SUBREDDITS:
                current += 1
                print(f"[{current}/{total_searches}] 搜索 r/{subreddit} - \"{keyword}\"")

                posts = search_subreddit(subreddit, keyword, limit=POSTS_PER_SEARCH)
                time.sleep(REQUEST_DELAY)

                for post in posts:
                    post_id = post["id"]
                    old = existing_data.get(post_id)
                    if old and len(old.get("top_comments", [])) >= 4:
                        continue

                    # 获取 top 评论
                    print(f"  -> 获取评论: {post['title'][:50]}...")
                    top_comments = fetch_top_comments(post["permalink"], top_n=TOP_COMMENTS)
                    time.sleep(REQUEST_DELAY)

                    all_rows.append({
                        "topic_idx": topic_idx,
                        "topic_name": topic_name,
                        "keyword": keyword,
                        "subreddit": post["subreddit"],
                        "title": post["title"],
                        "selftext": post["selftext"],
                        "score": post["score"],
                        "upvote_ratio": post["upvote_ratio"],
                        "num_comments": post["num_comments"],
                        "created_readable": utc_to_readable(post["created_utc"]),
                        "top_comments": top_comments,
                        "permalink": post["permalink"],
                    })

    except RateLimitAbort as e:
        abort_message = (
            f"\n⚠️ 连续 {MAX_CONSECUTIVE_RATE_LIMITS} 次被 Reddit 限流，爬取提前终止。\n"
            f"   终止位置: {e.message}\n"
            f"   已完成进度: {current}/{total_searches} 个搜索\n"
            f"   本次新采集: {len(all_rows)} 条"
        )
        print(abort_message)

    # 合并旧数据 + 新数据
    if all_rows or existing_data:
        merged = merge_rows(existing_data, all_rows)
        write_excel(merged, OUTPUT_FILE)
        if abort_message:
            print("\n📁 已将已爬取的数据保存到 Excel，可稍后继续补充。")
    else:
        print("\n⚠️ 未获取到任何数据，请检查网络连接或关键词设置。")

if __name__ == "__main__":
    main()