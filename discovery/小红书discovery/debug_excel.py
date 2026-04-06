"""Debug: 读取第一个 Excel 文件看 sheet 结构和数据"""
import os
import json
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

for root, dirs, files in os.walk(BASE_DIR):
    xlsx_files = [f for f in files if f.endswith(".xlsx")]
    if not xlsx_files:
        continue
    xlsx_path = os.path.join(root, xlsx_files[0])

    print(f"File: {xlsx_path[-50:]}")
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True)
        print(f"  Sheets: {wb.sheetnames}")

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            print(f"  [{sheet_name}] rows={len(rows)}")
            if rows:
                print(f"    Header: {rows[0]}")
                if len(rows) > 1:
                    print(f"    Row1: {rows[1]}")
                    if len(rows) > 2:
                        print(f"    Row2: {rows[2]}")
    except Exception as e:
        print(f"  ERROR: {e}")

    break

# Write result to file so we can read it
result_file = os.path.join(BASE_DIR, "debug_excel.json")
with open(result_file, "w", encoding="utf-8") as f:
    f.write("debug done")
