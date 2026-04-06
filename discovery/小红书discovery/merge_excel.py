"""
读取所有小红书 Excel 文件，合并 contents + comments，生成完整的 all_data_full.json
"""
import os
import json
import glob
import openpyxl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_FILE = os.path.join(BASE_DIR, "all_data_full.json")

def read_excel_sheets(filepath):
    wb = openpyxl.load_workbook(filepath, data_only=True)
    sheet_names_lower = {s.lower(): s for s in wb.sheetnames}
    contents_ws = wb[sheet_names_lower.get("contents")] if "contents" in sheet_names_lower else None
    comments_ws = wb[sheet_names_lower.get("comments")] if "comments" in sheet_names_lower else None
    return contents_ws, comments_ws

def sheet_to_records(sheet):
    if sheet is None:
        return []
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(h).strip() if h is not None else f"col_{i}" for i, h in enumerate(rows[0])]
    records = []
    for row in rows[1:]:
        if all(v is None for v in row):
            continue
        records.append({headers[i]: v for i, v in enumerate(row)})
    return records

def serialize_val(v):
    """确保所有值都是 JSON 可序列化的基本类型"""
    if isinstance(v, (int, float, str, bool)) or v is None:
        return v
    if isinstance(v, bytes):
        return v.decode("utf-8", errors="replace")
    return str(v)

def main():
    all_data = []
    stats = {}  # keyword -> {"posts": N, "comments": M}

    # 用 os.walk 扫描所有 xlsx（绕过目录名编码问题）
    for root, dirs, files in os.walk(BASE_DIR):
        xlsx_files = [f for f in files if f.endswith(".xlsx")]
        if not xlsx_files:
            continue

        # 从路径提取分类名（取直接父文件夹名）
        rel_root = os.path.relpath(root, BASE_DIR)
        if rel_root == ".":
            continue
        category_name = os.path.basename(rel_root)

        for xlsx_file in sorted(xlsx_files):
            xlsx_path = os.path.join(root, xlsx_file)
            keyword_name = os.path.splitext(xlsx_file)[0]

            try:
                contents_ws, comments_ws = read_excel_sheets(xlsx_path)
                contents_records = sheet_to_records(contents_ws)
                comments_records = sheet_to_records(comments_ws)
            except Exception as e:
                stats[keyword_name] = {"error": str(e), "posts": 0, "comments": 0}
                continue

            # 按 note_id 索引评论
            comments_by_note = {}
            for c in comments_records:
                nid = str(c.get("note_id") or c.get("note_id ") or "").strip()
                if nid:
                    # 序列化所有评论字段
                    serialized = {k: serialize_val(v) for k, v in c.items()}
                    comments_by_note.setdefault(nid, []).append(serialized)

            # 序列化所有帖子字段，附加 comments
            posts = []
            for post in contents_records:
                nid = str(post.get("note_id") or post.get("note_id ") or "").strip()
                serialized_post = {k: serialize_val(v) for k, v in post.items()}
                serialized_post["comments"] = comments_by_note.get(nid, [])
                posts.append(serialized_post)

            total_comments = sum(len(p["comments"]) for p in posts)
            stats[keyword_name] = {"posts": len(posts), "comments": total_comments}

            if posts:
                all_data.append({
                    "category": category_name,
                    "keyword": keyword_name,
                    "posts": posts
                })

    # 写入 JSON
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(all_data, f, ensure_ascii=False, indent=2)

    # 写入统计信息
    stats_file = os.path.join(BASE_DIR, "merge_stats.json")
    with open(stats_file, "w", encoding="utf-8") as f:
        json.dump(stats, f, ensure_ascii=False, indent=2)

    total_posts = sum(v["posts"] for v in stats.values() if "error" not in v)
    total_comments = sum(v["comments"] for v in stats.values() if "error" not in v)
    print(f"Done. {len(all_data)} categories, {total_posts} posts, {total_comments} comments")
    print(f"Output: {OUTPUT_FILE}")
    print(f"Stats: {stats_file}")

if __name__ == "__main__":
    main()
